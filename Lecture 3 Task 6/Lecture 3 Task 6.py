import os
import openpyxl

directory = os.path.dirname(os.path.abspath(__file__))
workbook = openpyxl.load_workbook(directory + "\\database.xlsx")
sheet = workbook.active
sheet.cell(1, 1).value = "Name"
sheet.cell(1, 2).value = "Job"
sheet.cell(1, 3).value = "Salary"
sheet.cell(1, 4).value = "ID"
print("Welcome!")
print("1- To create a new record: ")
print("2- To read an existing record: ")
print("3- To update an existing record: ")
print("4- To delete an existing record: ")
choice = int(input("Please enter an operation number: "))
if choice == 1:
    new_name = input("Enter name: ").capitalize()
    new_job = input("Enter job: ").capitalize()
    new_salary = input("Enter salary: ")
    current_row = sheet.max_row + 1
    sheet.cell(current_row, 1).value = new_name
    sheet.cell(current_row, 2).value = new_job
    sheet.cell(current_row, 3).value = new_salary
    sheet.cell(current_row, 4).value = f"{(sheet.max_row - 1):08d}"
elif choice == 2:
    is_found = False
    found_row = 0
    search = input("Enter any data to search for: ")
    for row in range(2, sheet.max_row + 1):
        if is_found:
            break
        for col in range(1, sheet.max_column + 1):
            if search == sheet.cell(row, col).value:
                is_found = True
                found_row = row
                break
    if is_found:
        print("Name: %s" % sheet.cell(found_row, 1).value)
        print("Job: %s" % sheet.cell(found_row, 2).value)
        print("Salary: %s" % sheet.cell(found_row, 3).value)
        print("ID: %s" % sheet.cell(found_row, 4).value)
    else:
        print("Data not found in any existing record")
elif choice == 3:
    search_id = input("Enter employee ID: ")
    row = 2
    while row <= sheet.max_row:
        if sheet.cell(row, 4).value == search_id:
            print("1- Name")
            print("2- Job")
            print("3- Salary")
            change = input("Choose what data to change: ")
            if change == "1" or change == "Name":
                sheet.cell(row, 1).value = input("Enter new name: ").capitalize()
                break
            elif change == "2" or change == "Job":
                sheet.cell(row, 2).value = input("Enter new job: ").capitalize()
                break
            elif change == "3" or change == "Salary":
                sheet.cell(row, 3).value = input("Enter new salary: ")
                break
            else:
                print("Not a valid option")
                row = 2
        else:
            row += 1
elif choice == 4:
    search_id = input("Enter employee ID: ")
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, 4).value == search_id:
            sheet.delete_rows(row)
            break
        else:
            continue
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 4).value = f"{(row - 1):08d}"
else:
    pass
workbook.save(directory + "\\database.xlsx")
