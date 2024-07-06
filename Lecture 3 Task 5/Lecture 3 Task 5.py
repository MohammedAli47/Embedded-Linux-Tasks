import os
import openpyxl

functions = []
directory = os.path.dirname(os.path.abspath(__file__))
with open(directory + "\\dummy.h", "r") as file:
    for line in file:
        functions.append(line[0:-1] if (line[0] != "#") else "")
while True:
    try:
        functions.remove("")
    except:
        break
workbook = openpyxl.load_workbook(directory + "\\database.xlsx")
sheet = workbook.active
for i in range(len(functions)):
    sheet.cell(i + 1, 1).value = functions[i]
    sheet.cell(i + 1, 2).value = "IDX" + f"{i:03d}"
workbook.save(directory + "\\database.xlsx")
